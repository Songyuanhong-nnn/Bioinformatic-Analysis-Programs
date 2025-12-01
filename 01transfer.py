import pandas as pd
import os
import glob
import random
import re
from openpyxl.styles import Font
from openpyxl import Workbook

def is_valid_template(df):
    """验证文件是否符合模板要求（保留初始核心逻辑：Gene列+GCF样本列）"""
    if "Gene" not in df.columns:
        return False, "缺少'Gene'列"
    samples = [col for col in df.columns if col.startswith("GCF_")]
    if not samples:
        return False, "未检测到GCF开头的样本列"
    return True, "符合模板"

def is_g_start_num_end(filename):
    """判断文件名是否符合“G开头且数字结尾”格式（新增筛选逻辑，不影响核心处理）"""
    pattern = r'^G.*\d+$'  # 严格匹配：以G开头，以1个及以上数字结尾
    file_basename = os.path.splitext(filename)[0]
    return bool(re.match(pattern, file_basename, re.IGNORECASE))  # 忽略大小写（G/g均匹配）

def sample_verify_base(final_matrix, df, file_basename, standard_genes, standard_samples):
    """基准版抽样检测：仅用当前文件数据验证，还原初始检测逻辑"""
    print("\n" + "="*40)
    print(f"🔍 检测【{file_basename}】矩阵正确性")
    print("="*40)
    
    if final_matrix.empty:
        print("⚠️  矩阵为空，跳过检测")
        return

    # 提取当前文件矩阵的基因和样本（仅当前文件数据）
    matrix_genes = final_matrix['基因簇名称'].tolist()
    matrix_samples = final_matrix.columns.tolist()[1:]
    if len(matrix_genes) == 0 or len(matrix_samples) == 0:
        print("⚠️  矩阵无有效数据，跳过检测")
        return

    # 初始版抽样逻辑：各取2-3个，避免过度抽样
    sample_genes = random.sample(matrix_genes, min(3, len(matrix_genes)))
    sample_samples = random.sample(matrix_samples, min(3, len(matrix_samples)))
    correct = 0
    total = len(sample_genes) * len(sample_samples)

    # 基于当前文件原始数据计算预期值（完全还原初始逻辑）
    for gene in sample_genes:
        for sample in sample_samples:
            # 获取矩阵实际值
            try:
                actual = final_matrix.loc[final_matrix['基因簇名称'] == gene, sample].iloc[0]
            except (IndexError, KeyError):
                print(f"❌ 基因'{gene}'_样本'{sample}'：矩阵中未找到")
                continue

            # 计算预期值（仅用当前文件原始数据）
            expected = 0
            # 筛选原始数据中该基因的行
            gene_rows = df[df['Gene'].str.strip() == gene.strip()]
            if not gene_rows.empty:
                # 检查该样本列是否有有效数据
                if sample in df.columns and pd.notna(gene_rows[sample].iloc[0]):
                    expected = 1

            # 结果判断
            if actual == expected:
                correct += 1
                print(f"✅ 基因'{gene}'_样本'{sample}'：实际{actual} = 预期{expected}")
            else:
                print(f"❌ 基因'{gene}'_样本'{sample}'：实际{actual} ≠ 预期{expected}")

    # 初始版检测结果输出
    print(f"\n📊 检测结果：{correct}/{total} 正确")
    if correct == total:
        print(f"🎉 【{file_basename}】矩阵验证通过")
    else:
        print(f"⚠️  【{file_basename}】矩阵存在错误，需检查")
    print("-"*40)

def main():
    """基准版主函数：完全还原单个文件独立处理逻辑，仅新增文件筛选和文件夹存储"""
    # 初始版路径设置（简洁，避免复杂路径逻辑）
    current_dir = os.path.dirname(os.path.abspath(__file__))
    standard_path = os.path.join(current_dir, "standard.csv")
    output_root = os.path.join(current_dir, "基因矩阵输出")  # 初始版输出根目录

    # 初始版文件夹初始化：仅创建，不清空（避免误删用户数据，还原初始安全逻辑）
    if not os.path.exists(output_root):
        os.makedirs(output_root)
        print(f"📂 创建输出根目录：{output_root}")

    # 1. 读取标准文件（仅用于标记颜色，不参与数据合并，还原初始用途）
    standard_genes = set()
    standard_samples = set()
    if os.path.exists(standard_path):
        try:
            std_df = pd.read_csv(standard_path, sep=",", low_memory=False)
            # 提取标准基因（去重去空）
            standard_genes = set([g.strip() for g in std_df['Gene'].dropna() if str(g).strip() != ""])
            # 提取标准样本（GCF开头）
            standard_samples = set([col.strip() for col in std_df.columns if col.startswith("GCF_")])
            print(f"\n📋 标准文件加载：{len(standard_genes)}个基因，{len(standard_samples)}个样本")
        except Exception as e:
            print(f"⚠️  标准文件读取错误：{str(e)}，将不标记标准行列")
    else:
        print(f"⚠️  未找到standard.csv，将不标记标准行列")

    # 2. 筛选待处理文件（新增G开头数字结尾筛选，不改变核心流程）
    csv_files = glob.glob(os.path.join(current_dir, "*.csv"))
    valid_files = []
    skipped_files = []
    for file in csv_files:
        filename = os.path.basename(file)
        # 排除标准文件和输出文件（初始版排除逻辑）
        if filename.lower() == "standard.csv" or "final" in filename.lower():
            continue
        # 新增筛选：仅保留G开头数字结尾的文件
        if is_g_start_num_end(filename):
            valid_files.append(file)
        else:
            skipped_files.append(filename)

    # 初始版文件筛选结果输出
    if skipped_files:
        print(f"\n⚠️  跳过非目标文件（需G开头数字结尾）：{', '.join(skipped_files)}")
    if len(valid_files) == 0:
        print(f"\n❌ 未找到可处理的CSV文件（需G开头数字结尾）")
        return
    print(f"\n📁 找到{len(valid_files)}个可处理文件：")
    for i, file in enumerate(valid_files, 1):
        print(f"   {i}. {os.path.basename(file)}")

    # 3. 单个文件独立处理（核心：完全还原初始单个文件处理逻辑，无多文件合并）
    for file in valid_files:
        filename = os.path.basename(file)
        file_basename = os.path.splitext(filename)[0]
        print(f"\n🚀 开始处理：{filename}")

        # 初始版文件读取逻辑
        try:
            df = pd.read_csv(file, sep=",", low_memory=False)
            print(f"📥 读取文件：{len(df)}行数据，{len(df.columns)}列")
        except Exception as e:
            print(f"❌ 读取{filename}错误：{str(e)}，跳过该文件")
            continue

        # 初始版文件验证
        is_valid, msg = is_valid_template(df)
        if not is_valid:
            print(f"❌ {filename}不符合模板：{msg}，跳过")
            continue

        # 4. 生成单个文件专属矩阵（完全初始逻辑：仅当前文件基因和样本）
        # 提取当前文件的基因（去重去空，排序保证一致性）
        file_genes = sorted(list(set([g.strip() for g in df['Gene'].dropna() if str(g).strip() != ""])))
        # 提取当前文件的样本列（仅GCF开头）
        file_samples = sorted([col.strip() for col in df.columns if col.startswith("GCF_")])
        print(f"📊 构建矩阵：{len(file_genes)}个基因 × {len(file_samples)}个样本")

        # 初始版矩阵初始化（全0）
        matrix = pd.DataFrame(0, index=range(len(file_genes)), columns=['基因簇名称'] + file_samples)
        matrix['基因簇名称'] = file_genes  # 填充基因列

        # 初始版矩阵填充逻辑（仅当前文件数据）
        for sample in file_samples:
            # 找到该样本列有有效数据的基因
            present_genes = [g.strip() for g in df[df[sample].notna()]['Gene'].dropna() if str(g).strip() != ""]
            # 更新矩阵值为1
            matrix.loc[matrix['基因簇名称'].isin(present_genes), sample] = 1
        print(f"✅ 矩阵填充完成")

        # 5. 初始版输出逻辑：按文件创建文件夹，存储CSV和Excel
        # 创建当前文件专属文件夹（新增功能，不影响核心处理）
        file_output_dir = os.path.join(output_root, f"{file_basename}_输出")
        if not os.path.exists(file_output_dir):
            os.makedirs(file_output_dir)

        # 保存CSV（初始版编码和格式）
        csv_path = os.path.join(file_output_dir, f"{file_basename}_matrix.csv")
        matrix.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"💾 CSV保存：{csv_path}")

        # 生成Excel（保留初始版格式逻辑：标准行列标红，交叉标蓝）
        wb = Workbook()
        ws = wb.active
        ws.title = f"{file_basename} Matrix"

        # 初始版Excel字体样式
        std_font = Font(bold=True, color="FF0000")  # 标准行列：红色加粗
        cross_font = Font(bold=True, color="0000FF")  # 标准交叉：蓝色加粗
        normal_font = Font(color="000000")  # 普通：黑色

        # 写入表头（初始版顺序）
        for col_idx, col_name in enumerate(matrix.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            # 表头样式：标准样本标红
            if col_name.strip() in standard_samples:
                ws.cell(row=1, column=col_idx).font = std_font
            else:
                ws.cell(row=1, column=col_idx).font = normal_font

        # 写入数据（初始版逐行处理）
        for row_idx, (_, row) in enumerate(matrix.iterrows(), 2):
            gene_name = row['基因簇名称'].strip()
            # 基因列样式：标准基因标红
            ws.cell(row=row_idx, column=1, value=gene_name)
            is_std_gene = gene_name in standard_genes
            if is_std_gene:
                ws.cell(row=row_idx, column=1).font = std_font
            else:
                ws.cell(row=row_idx, column=1).font = normal_font

            # 样本列数据和样式
            for col_idx, col_name in enumerate(matrix.columns[1:], 2):
                cell_value = row[col_name]
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
                # 标准交叉处标蓝（初始版逻辑）
                if is_std_gene and col_name.strip() in standard_samples:
                    ws.cell(row=row_idx, column=col_idx).font = cross_font
                else:
                    ws.cell(row=row_idx, column=col_idx).font = normal_font

        # 保存Excel
        excel_path = os.path.join(file_output_dir, f"{file_basename}_matrix.xlsx")
        wb.save(excel_path)
        print(f"💾 Excel保存：{excel_path}")

        # 6. 初始版抽样检测（仅当前文件）
        sample_verify_base(matrix, df, file_basename, standard_genes, standard_samples)

        # 初始版单个文件处理完成提示
        print(f"✅ 【{filename}】处理完成\n" + "-"*50)

    # 初始版最终总结
    print(f"\n🎯 所有文件处理完成！")
    print(f"📁 所有结果已保存至：{output_root}")
    print(f"💡 每个文件独立生成矩阵，数据与初始单个处理逻辑完全一致")

if __name__ == "__main__":
    main()