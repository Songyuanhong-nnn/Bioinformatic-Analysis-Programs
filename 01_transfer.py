import pandas as pd
import os
import glob
import random
from openpyxl.styles import Font
from openpyxl import Workbook

def is_valid_template(df):
    """验证文件是否符合模板要求（必须包含Gene列和GCF开头的样本列）"""
    if "Gene" not in df.columns:
        return False, "缺少'Gene'列"
    samples = [col for col in df.columns if col.startswith("GCF_")]
    if not samples:
        return False, "未检测到GCF开头的样本列"
    return True, "符合模板"

def sample_verify(final_matrix, all_input_data, standard_genes, standard_samples):
    """抽样检测：从最终矩阵中随机抽取样本，验证其0/1标记是否正确"""
    print("\n" + "="*50)
    print("🔍 开始抽样检测最终矩阵数据正确性...")
    print("="*50)
    
    if final_matrix.empty or not all_input_data:
        print("⚠️  矩阵为空或无输入数据，无法进行检测。")
        return

    all_genes = final_matrix['基因簇名称'].tolist()
    all_samples = final_matrix.columns.tolist()[1:]

    if not all_genes or len(all_samples) == 0:
        print("⚠️  矩阵中没有有效数据。")
        return

    standard_gene_samples = [g for g in all_genes if g in standard_genes]
    new_gene_samples = [g for g in all_genes if g not in standard_genes]
    sample_genes = []
    if standard_gene_samples:
        sample_genes.extend(random.sample(standard_gene_samples, min(3, len(standard_gene_samples))))
    if new_gene_samples:
        sample_genes.extend(random.sample(new_gene_samples, min(2, len(new_gene_samples))))
    sample_samples = random.sample(all_samples, min(5, len(all_samples)))

    sample_pool = [(g, s) for g in sample_genes for s in sample_samples]
    if not sample_pool:
        sample_pool = random.sample([(g, s) for g in all_genes for s in all_samples], min(10, len(all_genes)*len(all_samples)))

    correct_count = 0
    for gene, sample in sample_pool:
        try:
            actual_value = final_matrix.loc[final_matrix['基因簇名称'] == gene, sample].iloc[0]
        except (IndexError, KeyError):
            print(f"\n- 检测组合: 基因='{gene}', 样本='{sample}'")
            print(f"  结果: ❌ 错误 - 在最终矩阵中未找到该组合")
            continue

        expected_value = 0
        for df, file_samples in all_input_data:
            if sample in file_samples and gene in df["Gene"].str.strip().values:
                gene_row = df[df["Gene"].str.strip() == gene]
                if not gene_row.empty and pd.notna(gene_row[sample].iloc[0]):
                    expected_value = 1
                    break

        status = "✅ 正确" if expected_value == actual_value else "❌ 错误"
        if status == "✅ 正确":
            correct_count += 1
        
        gene_type = "标准基因" if gene in standard_genes else "新增基因"
        sample_type = "标准样本" if sample in standard_samples else "新增样本"
        print(f"\n- 检测组合: 基因='{gene}'（{gene_type}）, 样本='{sample}'（{sample_type}）")
        print(f"  预期值: {expected_value}")
        print(f"  实际值: {actual_value}")
        print(f"  结果: {status}")

    print("\n" + "-"*30)
    print(f"检测完成：{correct_count}/{len(sample_pool)} 条数据正确。")
    if correct_count == len(sample_pool):
        print("🎉 所有检测样本均正确，最终矩阵数据可信度高！")
    else:
        print("⚠️  部分检测样本不正确，请检查数据或程序逻辑。")
    print("-"*30)

def main():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    standard_filename = "standard.csv"
    standard_path = os.path.join(current_dir, standard_filename)
    output_csv = os.path.join(current_dir, "final_combined_matrix.csv")
    output_excel = os.path.join(current_dir, "final_combined_matrix.xlsx")

    # 强制删除旧文件，避免缓存干扰
    for f in [output_csv, output_excel]:
        if os.path.exists(f):
            os.remove(f)
            print(f"🗑️  已强制删除旧文件：{os.path.basename(f)}")

    # 1. 读取标准文件
    if not os.path.exists(standard_path):
        print(f"❌ 未找到标准表格：{standard_filename}，请放在程序目录下")
        return

    try:
        std_df = pd.read_csv(standard_path, sep=",", header=0, low_memory=False)
        raw_standard_genes = std_df["Gene"].dropna().tolist()
        standard_genes = set([g.strip() for g in raw_standard_genes if pd.notna(g)])
        standard_samples = set([col.strip() for col in std_df.columns if col.startswith("GCF_")])
        
        print(f"📋 从标准文件识别到：")
        print(f"   - 去重后标准基因数：{len(standard_genes)} 个")
        print(f"   - 标准样本数：{len(standard_samples)} 个")
        
        if not standard_genes or not standard_samples:
            return
    except Exception as e:
        print(f"❌ 读取{standard_filename}失败：{str(e)}")
        return

    # 2. 扫描待处理CSV
    all_csv_files = glob.glob(os.path.join(current_dir, "*.csv"))
    valid_files = []
    for file in all_csv_files:
        filename = os.path.basename(file)
        if filename.lower() == standard_filename.lower() or filename == os.path.basename(output_csv):
            continue
        valid_files.append(file)

    if not valid_files:
        print("⚠️ 未检测到可处理的CSV文件！")
        return

    print(f"\n📁 检测到 {len(valid_files)} 个可处理文件，开始构建并集...")
    for f in valid_files:
        print(f"  - {os.path.basename(f)}")
    print("-" * 30 + "\n")

    # 3. 收集所有基因和样本（并集）
    all_input_data = []
    all_genes_set = set(standard_genes)
    all_samples_set = set(standard_samples)

    for csv_file in valid_files:
        try:
            df = pd.read_csv(csv_file, sep=",", header=0, low_memory=False)
            valid, msg = is_valid_template(df)
            if not valid:
                print(f"⚠️ 跳过 {os.path.basename(csv_file)}：{msg}\n")
                continue

            file_genes = [g.strip() for g in df["Gene"].dropna().tolist()]
            unique_file_genes = set(file_genes)
            file_samples = [col.strip() for col in df.columns if col.startswith("GCF_")]
            unique_file_samples = set(file_samples)

            all_genes_set.update(unique_file_genes)
            all_samples_set.update(unique_file_samples)
            all_input_data.append( (df, unique_file_samples) )

            print(f"✅ 处理 {os.path.basename(csv_file)}：")
            print(f"   - 新增基因数：{len(unique_file_genes - standard_genes)} 个")
            print(f"   - 新增样本数：{len(unique_file_samples - standard_samples)} 个")

        except Exception as e:
            print(f"❌ 处理 {os.path.basename(csv_file)} 出错：{str(e)}\n")

    # 4. 最终行列顺序
    all_genes = list(standard_genes) + sorted(list(all_genes_set - standard_genes))
    all_samples = list(standard_samples) + sorted(list(all_samples_set - standard_samples))

    print(f"\n📊 最终并集结果：")
    print(f"   - 总基因数：{len(all_genes)} 个（标准{len(standard_genes)} + 新增{len(all_genes)-len(standard_genes)}）")
    print(f"   - 总样本数：{len(all_samples)} 个（标准{len(standard_samples)} + 新增{len(all_samples)-len(standard_samples)}）")

    # 5. 填充矩阵
    final_matrix = pd.DataFrame(0, index=all_genes, columns=all_samples)
    print(f"\n📥 正在填充矩阵（{len(all_genes)}行 × {len(all_samples)}列）...")
    for df, file_samples in all_input_data:
        for sample in file_samples:
            if sample not in final_matrix.columns:
                continue
            present_genes = [g.strip() for g in df[df[sample].notna()]["Gene"].dropna().tolist()]
            final_matrix.loc[present_genes, sample] = 1

    # 6. 格式化并保存
    final_matrix.reset_index(inplace=True)
    final_matrix.rename(columns={"index": "基因簇名称"}, inplace=True)

    # 保存CSV
    final_matrix.to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"✅ CSV文件已保存：{output_csv}")

    # 生成Excel并应用最终格式
    print(f"✅ 正在应用最终格式...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Matrix"

    # 定义字体样式
    standard_label_font = Font(bold=True, color="FF0000")  # 标准基因名和样本名：加粗红色
    default_font = Font(color="000000")                    # 默认：黑色
    cross_highlight_font = Font(bold=True, color="0000FF") # 标准基因与标准样本交叉处：加粗蓝色

    # 写入表头并设置格式
    for col_idx, col_name in enumerate(final_matrix.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
        if col_name in standard_samples:
            ws.cell(row=1, column=col_idx).font = standard_label_font
        else:
            ws.cell(row=1, column=col_idx).font = default_font

    # 写入数据并设置格式
    for row_idx, (_, row) in enumerate(final_matrix.iterrows(), 2):
        gene_name = str(row["基因簇名称"]).strip() if pd.notna(row["基因簇名称"]) else ""
        
        # 设置基因名格式
        ws.cell(row=row_idx, column=1, value=gene_name)
        is_standard_gene_row = gene_name in standard_genes
        if is_standard_gene_row:
            ws.cell(row=row_idx, column=1).font = standard_label_font
        else:
            ws.cell(row=row_idx, column=1).font = default_font

        # 设置数值格式
        for col_idx, col_name in enumerate(final_matrix.columns[1:], 2):
            cell_value = row[col_name]
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # 判断是否为标准基因和标准样本的交叉点
            if is_standard_gene_row and (col_name in standard_samples):
                ws.cell(row=row_idx, column=col_idx).font = cross_highlight_font
            else:
                ws.cell(row=row_idx, column=col_idx).font = default_font

    # 保存Excel
    wb.save(output_excel)
    print(f"✅ Excel文件已保存：{output_excel}")

    # 最终格式说明
    print(f"\n🎨 最终格式汇总（完全按需求）：")
    print(f"   1. 【红色加粗】'基因簇名称'列中，存在于 standard.csv 的基因名。")
    print(f"   2. 【红色加粗】表头中，存在于 standard.csv 的样本名。")
    print(f"   3. 【蓝色加粗】同时满足以下两个条件的单元格数值：")
    print(f"      - 所在行的基因是标准基因（红色）。")
    print(f"      - 所在列的样本是标准样本（红色）。")
    print(f"   4. 【黑色】其他所有单元格（包括数值和非标准的行列名）。")

    # 7. 抽样检测
    sample_verify(final_matrix, all_input_data, standard_genes, standard_samples)

if __name__ == "__main__":
    main()
