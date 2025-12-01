#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=====================================================================
🚀 智能菌株危害等级评分+高毒基因筛选工具（最终纯Python版）
=====================================================================
✅ WSL路径自动转换 | 一键默认路径 | 错误提示清晰
✅ 全自动依赖安装 | 教程级引导 | 高容错重填
=====================================================================
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import matplotlib.pyplot as plt
import seaborn as sns
import os
import sys
import subprocess
import json
from datetime import datetime

# -------------------------- 全局配置 --------------------------
BACKUP_FILE = ".analysis_backup.json"
PRESET_CONFIG = {
    "name": "预设评分体系（致病关联40%+毒力基因35%+爆发关联25%）",
    "indicators": ["致病关联分", "毒力基因携带分", "爆发关联分"],
    "weights": [0.4, 0.35, 0.25],
    "threshold": 70.0
}

# 设置中文字体
def setup_font():
    try:
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans', 'WenQuanYi Zen Hei']
        plt.rcParams['axes.unicode_minus'] = False
    except:
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False

setup_font()

# -------------------------- 工具函数 --------------------------
def print_title(msg):
    print(f"\n{'='*80}")
    print(f"📌 {msg}")
    print(f"{'='*80}")

def print_success(msg):
    print(f"✅ {msg}")

def print_warning(msg):
    print(f"⚠️  {msg}")

def print_error(msg):
    print(f"❌ {msg}")

def print_example(msg):
    print(f"📋 示例：{msg}")

def print_help(msg):
    print(f"💡 提示：{msg}")

def confirm_operation(msg):
    while True:
        res = input(f"{msg}（默认y，输入n取消）：").strip().lower()
        if res in ["y", "n", ""]:
            return res == "y" or res == ""
        print_error("输入错误！请输入 y 或 n")

def load_backup():
    if os.path.exists(BACKUP_FILE):
        try:
            with open(BACKUP_FILE, "r", encoding="utf-8") as f:
                backup = json.load(f)
            if confirm_operation(f"检测到上次未完成的分析进度，是否恢复？"):
                print_success("已恢复上次进度！")
                return backup
            else:
                os.remove(BACKUP_FILE)
                return None
        except:
            os.remove(BACKUP_FILE)
            return None
    return None

def save_backup(data):
    data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(BACKUP_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def delete_backup():
    if os.path.exists(BACKUP_FILE):
        os.remove(BACKUP_FILE)

# -------------------------- WSL 路径智能转换 --------------------------
def convert_windows_to_wsl_path(path):
    """将 Windows 路径（D:\\...）转换为 WSL 路径（/mnt/d/...）"""
    if not path:
        return path
    path = path.replace("\\", "/")
    if len(path) >= 2 and path[1] == ":":
        drive = path[0].lower()
        wsl_path = f"/mnt/{drive}/{path[2:]}"
        wsl_path = os.path.normpath(wsl_path)
        return wsl_path
    return path

def check_path_similarity(input_path, target_dir):
    input_path = input_path.lower().replace("\\", "/").rstrip("/")
    target_dir = target_dir.lower().replace("\\", "/").rstrip("/")
    input_parts = input_path.split("/")
    target_parts = target_dir.split("/")
    match_count = 0
    for i in range(min(len(input_parts), len(target_parts))):
        if input_parts[i] == target_parts[i]:
            match_count += 1
        else:
            break
    if match_count >= len(target_parts) - 1 and match_count > 0:
        print_warning(f"检测到你输入的路径接近：{target_dir}")
        print_help(f"是否要使用正确路径？直接输入：{target_dir}/gene_presence_absence.Rtab")

# -------------------------- 全自动依赖管理 --------------------------
def check_and_install_dependencies():
    required_packages = {
        "pandas": "数据处理",
        "openpyxl": "Excel模板生成",
        "matplotlib": "可视化图表",
        "seaborn": "美化图表"
    }
    missing_packages = []

    for pkg, desc in required_packages.items():
        try:
            __import__(pkg)
            print_success(f"{pkg}（{desc}）已安装")
        except ImportError:
            missing_packages.append((pkg, desc))
            print_warning(f"{pkg}（{desc}）未安装，将自动安装")

    if missing_packages:
        print_title("开始安装缺失依赖（请耐心等待）")
        try:
            install_cmd = [sys.executable, "-m", "pip", "install", "--user", "--upgrade"]
            install_cmd += [pkg for pkg, _ in missing_packages]
            subprocess.check_call(install_cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            print_success("所有依赖安装完成！脚本将自动重启...")
            os.execv(sys.executable, [sys.executable] + sys.argv)
        except Exception as e:
            print_error(f"依赖安装失败！错误原因：{str(e)[:100]}")
            print_help("请手动运行以下命令安装依赖，然后重新启动脚本：")
            print(f"pip3 install --user {' '.join([pkg for pkg, _ in missing_packages])}")
            sys.exit(1)

# -------------------------- 智能输入模块 --------------------------
def select_score_system(backup):
    if backup and "score_system" in backup:
        return backup["score_system"], backup["indicators"], backup["weights"], backup["threshold"]
    
    print_title("1. 选择评分体系（推荐直接选预设）")
    print("请选择评分体系类型：")
    print("  1. 预设评分体系（你的需求：致病关联40% + 毒力基因35% + 爆发关联25%，阈值70分）")
    print("  2. 自定义评分体系（可修改指标、权重、阈值）")
    
    while True:
        choice = input("输入数字1或2（默认1）：").strip()
        if choice == "" or choice == "1":
            print_success(f"已选择：{PRESET_CONFIG['name']}")
            print(f"  - 评分指标：{', '.join(PRESET_CONFIG['indicators'])}")
            print(f"  - 指标权重：{', '.join([f'{w:.2f}' for w in PRESET_CONFIG['weights']])}")
            print(f"  - 高危害阈值：{PRESET_CONFIG['threshold']}分（>该值为高毒）")
            
            if confirm_operation("是否确认使用该预设体系？"):
                save_backup({
                    "score_system": "preset",
                    "indicators": PRESET_CONFIG["indicators"],
                    "weights": PRESET_CONFIG["weights"],
                    "threshold": PRESET_CONFIG["threshold"]
                })
                return "preset", PRESET_CONFIG["indicators"], PRESET_CONFIG["weights"], PRESET_CONFIG["threshold"]
        elif choice == "2":
            print_title("自定义评分体系配置")
            indicators = define_indicators(None)
            weights = define_weights(indicators, None)
            threshold = define_threshold(None)
            
            print_title("自定义配置确认")
            print(f"  - 评分指标：{', '.join(indicators)}")
            print(f"  - 指标权重：{', '.join([f'{w:.2f}' for w in weights])}")
            print(f"  - 高危害阈值：{threshold}分")
            
            if confirm_operation("是否确认使用该自定义体系？"):
                save_backup({
                    "score_system": "custom",
                    "indicators": indicators,
                    "weights": weights,
                    "threshold": threshold
                })
                return "custom", indicators, weights, threshold
        else:
            print_error("输入错误！请输入1或2")

def define_indicators(backup):
    if backup and "indicators" in backup:
        return backup["indicators"]
    
    print_title("自定义：1. 定义评分指标")
    print_help("评分指标是量化菌株危害的维度，每个指标后续需填写0-100分（100分=危害最高）")
    print_example("致病关联分,毒力基因携带分,爆发关联分,耐药基因携带分")
    
    while True:
        input_str = input("请输入评分指标（默认：致病关联分,毒力基因携带分,爆发关联分）：").strip()
        if input_str == "":
            indicators = ["致病关联分", "毒力基因携带分", "爆发关联分"]
        else:
            indicators = [ind.strip() for ind in input_str.split(",") if ind.strip()]
        
        if len(indicators) < 1:
            print_error("错误：至少需要1个评分指标！")
            continue
        
        print_success(f"已定义 {len(indicators)} 个指标：")
        for i, ind in enumerate(indicators, 1):
            print(f"  1. {ind}")
        
        if confirm_operation("是否确认？"):
            return indicators

def define_weights(indicators, backup):
    if backup and "weights" in backup:
        return backup["weights"]
    
    print_title(f"自定义：2. 定义指标权重（共{len(indicators)}个指标）")
    print_help(f"权重总和必须=1（允许±0.01误差），按以下顺序输入：")
    for i, ind in enumerate(indicators, 1):
        print(f"  {i}. {ind}")
    print_example("0.4,0.35,0.25（总和=1）")
    
    while True:
        input_str = input("请输入权重（默认：0.4,0.35,0.25）：").strip()
        if input_str == "":
            weights = [0.4, 0.35, 0.25] if len(indicators) == 3 else [1.0/len(indicators)]*len(indicators)
        else:
            try:
                weights = [float(w.strip()) for w in input_str.split(",") if w.strip()]
            except ValueError:
                print_error("错误：权重必须是数字（如0.4、0.35）！")
                continue
        
        if len(weights) != len(indicators):
            print_error(f"错误：权重数量（{len(weights)}）与指标数量（{len(indicators)}）不一致！")
            continue
        
        weight_sum = sum(weights)
        if not (0.99 <= weight_sum <= 1.01):
            print_error(f"错误：权重总和={weight_sum:.3f}，必须接近1！")
            continue
        
        print_success(f"权重配置完成（总和={weight_sum:.3f}）：")
        for ind, w in zip(indicators, weights):
            print(f"  {ind}：{w:.3f}")
        
        if confirm_operation("是否确认？"):
            return weights

def define_threshold(backup):
    if backup and "threshold" in backup:
        return backup["threshold"]
    
    print_title("自定义：3. 定义高危害阈值")
    print_help("得分>阈值→高危害，得分≤阈值→低危害（范围0-100）")
    print_example("70（>70分为高毒）")
    
    while True:
        input_str = input("请输入阈值（0-100，默认70）：").strip()
        if input_str == "":
            threshold = 70.0
        else:
            try:
                threshold = float(input_str)
            except ValueError:
                print_error("错误：阈值必须是数字（如70、65）！")
                continue
        
        if not (0 <= threshold <= 100):
            print_error("错误：阈值必须在0-100之间！")
            continue
        
        print_success(f"已设置高危害阈值：{threshold:.1f}分")
        if confirm_operation("是否确认？"):
            return threshold

def get_rtab_path(backup):
    current_dir = os.path.abspath(".")
    target_filename = "gene_presence_absence.Rtab"
    current_dir_file = os.path.join(current_dir, target_filename)
    has_current_file = os.path.exists(current_dir_file)
    
    if backup and "rtab_path" in backup:
        if os.path.exists(backup["rtab_path"]):
            print_success(f"已恢复Rtab文件路径：{backup['rtab_path']}")
            return backup["rtab_path"]
        else:
            print_warning("备份的Rtab路径不存在，需重新输入")
    
    print_title("2. 输入核心文件路径（gene_presence_absence.Rtab）")
    if has_current_file:
        print_help(f"✅ 检测到脚本所在目录已存在 {target_filename}，直接按回车即可使用！")
    else:
        print_help("当前目录未找到文件，请输入完整路径")
    
    print("📋 支持的路径格式（WSL专用）：")
    print(f"  1. 相对路径（同一目录）：直接回车 或 输入 {target_filename}")
    print(f"  2. WSL绝对路径：/mnt/d/WSL/disk/projects/VP1/roary_output_final_1762658265/{target_filename}")
    print(f"  3. Windows路径（自动转换）：D:\\WSL\\disk\\projects\\VP1\\roary_output_final_1762658265\\{target_filename}")
    print_example(f"正确示例（WSL）：/mnt/d/WSL/disk/projects/VP1/roary_output_final_1762658265/{target_filename}")
    
    while True:
        rtab_path = input(f"请输入文件路径（默认：{target_filename}）：").strip()
        
        if rtab_path == "":
            if has_current_file:
                print_success(f"已选择当前目录的文件：{current_dir_file}")
                backup_data = load_backup() or {}
                backup_data["rtab_path"] = current_dir_file
                save_backup(backup_data)
                return current_dir_file
            else:
                print_error("错误：当前目录未找到 gene_presence_absence.Rtab，请手动输入路径")
                continue
        
        rtab_path = convert_windows_to_wsl_path(rtab_path)
        print_help(f"已自动转换路径为：{rtab_path}")
        
        if os.path.exists(rtab_path):
            try:
                df_test = pd.read_csv(rtab_path, sep="\t", index_col=0, nrows=5, low_memory=False)
                gene_count = pd.read_csv(rtab_path, sep="\t", index_col=0).shape[0]
                strain_count = pd.read_csv(rtab_path, sep="\t", index_col=0).shape[1]
                
                if strain_count < 2:
                    print_error("错误：文件结构无效，至少需要2个菌株！")
                    continue
                
                values = df_test.values.flatten()
                if not all(v in [0, 1] for v in values if pd.notna(v)):
                    print_warning("警告：文件中存在非0/1的值，可能不是Roary标准输出")
                    if not confirm_operation("是否继续使用该文件？"):
                        continue
                
                print_success("文件校验通过！")
                print(f"  - 基因数量：{gene_count:,} 个")
                print(f"  - 菌株数量：{strain_count:,} 个")
                
                backup_data = load_backup() or {}
                backup_data["rtab_path"] = rtab_path
                save_backup(backup_data)
                
                return rtab_path
            except Exception as e:
                print_error(f"文件格式错误！错误原因：{str(e)[:100]}")
                print_help("请确认文件是Roary输出的.Rtab格式（制表符分隔，首列基因名）")
                continue
        else:
            print_error(f"未找到文件：{rtab_path}")
            check_path_similarity(rtab_path, current_dir)
            print_help("可能的错误原因：")
            print("  1. 目录拼写错误（如 roary_output_final_1762658265 输错）")
            print("  2. 盘符错误（WSL中D盘是 /mnt/d，不是 /mnt/D）")
            print("  3. 文件名错误（必须是 gene_presence_absence.Rtab）")

# -------------------------- 模板生成+校验 --------------------------
def generate_excel_template(rtab_path, indicators, threshold, backup):
    if backup and "template_path" in backup and os.path.exists(backup["template_path"]):
        template_path = backup["template_path"]
        print_success(f"已恢复Excel模板：{template_path}")
        if confirm_operation("是否直接使用该模板？"):
            return template_path
    
    print_title("3. 生成Excel评分模板（含自动计算公式）")
    df_roary = pd.read_csv(rtab_path, sep="\t", index_col=0, low_memory=False)
    all_strains = df_roary.columns.tolist()
    print_success(f"已提取 {len(all_strains):,} 个菌株名称")
    
    template_data = {"菌株完整名称": all_strains}
    for ind in indicators:
        template_data[ind] = [None] * len(all_strains)
    template_data["总得分"] = [None] * len(all_strains)
    template_data["危害等级"] = [None] * len(all_strains)
    
    df_template = pd.DataFrame(template_data)
    rtab_dir = os.path.dirname(rtab_path)
    template_filename = f"菌株危害评分模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
    template_path = os.path.join(rtab_dir, template_filename)
    
    df_template.to_excel(template_path, index=False, engine="openpyxl")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
    
    start_row = 2
    global current_weights
    for row in range(start_row, ws.max_row + 1):
        formula_parts = [f"{chr(65 + i)}{row}*{current_weights[i]}" for i in range(len(indicators))]
        score_formula = f"=SUM({'+'.join(formula_parts)})"
        ws.cell(row=row, column=len(indicators) + 2).value = score_formula
        
        level_col = len(indicators) + 3
        level_formula = f'=IF({chr(65 + len(indicators) + 1)}{row}>{threshold}, "高", "低")'
        ws.cell(row=row, column=level_col).value = level_formula
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)
    
    wb.save(template_path)
    print_success(f"Excel模板已生成：{template_path}")
    
    print_title("Excel模板使用教程")
    print("📝 填写步骤：")
    print("  1. 用Excel打开模板文件（在Windows中找到该文件，双击打开）")
    print("  2. 在「致病关联分」「毒力基因携带分」「爆发关联分」列填写0-100分（100分=危害最高）")
    print("  3. 请勿修改「菌株完整名称」列，「总得分」和「危害等级」会自动计算！")
    print("⚠️  注意：不要留空，仅填写数字（如50、85、100）")

def wait_for_template_fill(template_path, indicators):
    print_title("4. 等待填写Excel模板")
    print_help(f"模板文件路径：{template_path}")
    print_help("填写完成后，回到终端输入 'yes' 继续分析，输入 'check' 校验填写是否正确")
    
    while True:
        confirm = input("请输入 'yes' 继续或 'check' 校验：").strip().lower()
        
        if confirm == "check":
            try:
                df_filled = pd.read_excel(template_path)
                missing_cols = [col for col in indicators if col not in df_filled.columns]
                if missing_cols:
                    print_error(f"模板缺少指标列：{', '.join(missing_cols)}")
                    continue
                
                empty_cells = {}
                for ind in indicators:
                    empty_count = df_filled[ind].isnull().sum()
                    if empty_count > 0:
                        empty_cells[ind] = empty_count
                
                if empty_cells:
                    print_warning("以下指标列存在未填写的单元格：")
                    for ind, cnt in empty_cells.items():
                        print(f"  - {ind}：{cnt} 个")
                    print_help("请补全所有单元格后再继续！")
                else:
                    print_success("✅ 填写完整性校验通过！")
                    invalid_scores = {}
                    for ind in indicators:
                        invalid = df_filled[(df_filled[ind] < 0) | (df_filled[ind] > 100)][ind].count()
                        if invalid > 0:
                            invalid_scores[ind] = invalid
                    
                    if invalid_scores:
                        print_warning("以下指标列存在无效得分（必须是0-100的数字）：")
                        for ind, cnt in invalid_scores.items():
                            print(f"  - {ind}：{cnt} 个")
                        print_help("请修改为0-100之间的数字！")
                    else:
                        print_success("✅ 得分范围校验通过！可以输入 'yes' 开始分析了")
            except Exception as e:
                print_error(f"校验失败：{str(e)[:100]}")
                print_help("可能是模板文件被修改或损坏，请重新生成模板（重启脚本）")
        elif confirm == "yes":
            try:
                df_filled = pd.read_excel(template_path)
                missing_cols = [col for col in indicators if col not in df_filled.columns]
                if missing_cols:
                    print_error(f"错误：模板缺少指标列 {', '.join(missing_cols)}")
                    continue
                for ind in indicators:
                    if df_filled[ind].isnull().any():
                        print_error(f"错误：{ind} 列有未填写的单元格！")
                        print_help("请补全后再输入 'yes'")
                        continue
                    if (df_filled[ind] < 0).any() or (df_filled[ind] > 100).any():
                        print_error(f"错误：{ind} 列有超出0-100的得分！")
                        print_help("请修改为0-100之间的数字后再输入 'yes'")
                        continue
                if "总得分" not in df_filled.columns or "危害等级" not in df_filled.columns:
                    print_error("错误：模板缺少「总得分」或「危害等级」列！")
                    continue
                print_success("模板填写合格，开始分析！")
                return df_filled
            except Exception as e:
                print_error(f"读取模板失败：{str(e)[:100]}")
        else:
            print_error("输入错误！只能输入 'yes' 或 'check'")

# -------------------------- 核心分析模块 --------------------------
def analyze_data(rtab_path, df_filled, indicators, weights, threshold):
    print_title("5. 开始智能分析")
    
    print("🔍 提取高/低毒菌株分组...")
    high_strains = df_filled[df_filled["危害等级"] == "高"]["菌株完整名称"].unique().tolist()
    low_strains = df_filled[df_filled["危害等级"] == "低"]["菌株完整名称"].unique().tolist()
    
    if len(high_strains) == 0:
        print_error("分析失败：未识别到高危害菌株！")
        print_help("请调整评分标准，确保有菌株总得分>70分")
        sys.exit(1)
    if len(low_strains) == 0:
        print_error("分析失败：未识别到低危害菌株！")
        print_help("请调整评分标准，确保有菌株总得分≤70分")
        sys.exit(1)
    
    print_success(f"菌株分组结果：")
    print(f"  - 高危害菌株：{len(high_strains):,} 株")
    print(f"  - 低危害菌株：{len(low_strains):,} 株")
    
    df_roary = pd.read_csv(rtab_path, sep="\t", index_col=0, low_memory=False)
    rtab_strains = df_roary.columns.tolist()
    
    high_strains_valid = [s for s in high_strains if s in rtab_strains]
    low_strains_valid = [s for s in low_strains if s in rtab_strains]
    
    if len(high_strains_valid) != len(high_strains):
        print_warning(f"⚠️  {len(high_strains)-len(high_strains_valid)} 个高毒菌株名在Rtab文件中未找到，已自动过滤")
    if len(low_strains_valid) != len(low_strains):
        print_warning(f"⚠️  {len(low_strains)-len(low_strains_valid)} 个低毒菌株名在Rtab文件中未找到，已自动过滤")
    
    print("🔍 筛选高毒专属候选基因（仅高危害菌株有，低危害菌株无的基因）...")
    high_vir_genes = []
    total_genes = len(df_roary.index)
    
    for i, gene in enumerate(df_roary.index, 1):
        if i % 1000 == 0:
            print(f"  进度：{i}/{total_genes} 个基因（{i/total_genes*100:.1f}%）")
        
        high_presence = df_roary.loc[gene, high_strains_valid].all()  # 所有高毒菌株都有这个基因
        low_absence = (df_roary.loc[gene, low_strains_valid] == 0).all()  # 所有低毒菌株都没有这个基因
        
        if high_presence and low_absence:
            high_vir_genes.append(gene)
    
    print_success(f"基因筛选完成！共找到 {len(high_vir_genes):,} 个高毒专属候选基因")
    
    rtab_dir = os.path.dirname(rtab_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 保存菌株分组文件
    group_path = os.path.join(rtab_dir, f"高毒低毒菌株分组_{timestamp}.txt")
    group_data = pd.DataFrame({
        "菌株完整名称": high_strains_valid + low_strains_valid,
        "危害等级": ["高"]*len(high_strains_valid) + ["低"]*len(low_strains_valid)
    })
    group_data.to_csv(group_path, sep="\t", index=False, header=True)
    print_success(f"菌株分组文件已保存：{group_path}")
    
    # 保存候选基因文件
    genes_path = os.path.join(rtab_dir, f"高毒专属候选基因_{timestamp}.txt")
    pd.DataFrame({"高毒专属候选基因": high_vir_genes}).to_csv(genes_path, index=False, header=True)
    print_success(f"候选基因文件已保存：{genes_path}")
    
    # 生成可视化报告
    print("📊 生成可视化分析报告...")
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
    
    # 图1：总得分分布
    sns.histplot(df_filled["总得分"], bins=30, ax=ax1, color="#4ecdc4", edgecolor="black", alpha=0.8)
    ax1.axvline(threshold, color="#ff6b6b", linestyle="--", linewidth=3, label=f"高/低毒阈值：{threshold}分")
    ax1.set_title("菌株危害总得分分布", fontsize=14, fontweight="bold")
    ax1.set_xlabel("总得分")
    ax1.set_ylabel("菌株数量")
    ax1.legend()
    ax1.grid(alpha=0.3)
    
    # 图2：高/低毒菌株占比
    counts = [len(high_strains), len(low_strains)]
    labels = [f"高危害菌株\n{len(high_strains):,}株", f"低危害菌株\n{len(low_strains):,}株"]
    colors = ["#ff6b6b", "#4ecdc4"]
    ax2.pie(counts, labels=labels, colors=colors, autopct="%1.1f%%", startangle=90)
    ax2.set_title("高/低危害菌株占比", fontsize=14, fontweight="bold")
    
    # 图3：各指标得分分布
    df_melt = pd.melt(df_filled, id_vars=["菌株完整名称", "总得分", "危害等级"], value_vars=indicators, var_name="评分指标", value_name="得分")
    sns.boxplot(x="评分指标", y="得分", hue="危害等级", data=df_melt, ax=ax3, palette=colors, alpha=0.8)
    ax3.set_title("各评分指标得分分布", fontsize=14, fontweight="bold")
    ax3.set_xlabel("评分指标")
    ax3.set_ylabel("得分")
    ax3.tick_params(axis='x', rotation=45)
    
    # 图4：基因筛选统计
    gene_stats = pd.DataFrame({
        "统计项": ["泛基因组总基因数", "高毒专属候选基因数", "候选基因占比"],
        "数值": [f"{total_genes:,}", f"{len(high_vir_genes):,}", f"{len(high_vir_genes)/total_genes*100:.2f}%"]
    })
    ax4.axis("off")
    ax4.table(cellText=gene_stats.values, colLabels=gene_stats.columns, cellLoc="center", loc="center", colWidths=[0.4, 0.6])
    ax4.set_title("基因筛选结果统计", fontsize=14, fontweight="bold")
    
    # 保存可视化图
    viz_path = os.path.join(rtab_dir, f"危害评分分析可视化报告_{timestamp}.png")
    plt.tight_layout()
    plt.savefig(viz_path, dpi=300, bbox_inches="tight", facecolor="white")
    plt.close()
    print_success(f"可视化分析报告已保存：{viz_path}")
    
    return group_path, genes_path, viz_path

# -------------------------- 主流程 --------------------------
def main():
    print_title("🚀 智能菌株危害等级评分+高毒基因筛选工具（最终纯Python版）")
    print("✅ 无任何多余命令 | 纯Python代码 | 直接运行即可")
    print("✅ 支持WSL/Windows路径 | 全自动依赖安装 | 详细教程引导")
    print("="*80)
    
    # 加载备份进度
    backup = load_backup()
    
    # 安装缺失依赖
    check_and_install_dependencies()
    
    # 选择评分体系
    score_system, indicators, weights, threshold = select_score_system(backup)
    global current_weights
    current_weights = weights
    
    # 输入Rtab文件路径
    rtab_path = get_rtab_path(backup)
    
    # 生成Excel评分模板
    template_path = generate_excel_template(rtab_path, indicators, threshold, backup)
    
    # 等待填写模板并校验
    df_filled = wait_for_template_fill(template_path, indicators)
    
    # 核心分析
    group_path, genes_path, viz_path = analyze_data(rtab_path, df_filled, indicators, weights, threshold)
    
    # 删除备份（分析完成）
    delete_backup()
    
    # 输出结果总结
    print_title("🎉 所有分析流程全部完成！")
    print("📁 生成的结果文件清单（都在当前目录）：")
    print(f"  1. Excel评分模板：{template_path}（你填写的原始评分数据）")
    print(f"  2. 菌株分组文件：{group_path}（高/低危害菌株列表）")
    print(f"  3. 高毒专属候选基因：{genes_path}（仅高毒菌株有，低毒菌株无的基因）")
    print(f"  4. 可视化分析报告：{viz_path}（4张图表，直接用于论文/汇报）")
    print("\n💡 下一步操作建议：")
    print("  1. 用 VirulenceFinder（https://cge.food.dtu.dk/services/VirulenceFinder/）验证候选基因的毒力功能")
    print("  2. 用 VFDB（http://www.mgc.ac.cn/VFs/）数据库注释候选基因的功能")
    print("  3. 可视化报告可直接插入论文或项目汇报PPT")
    print("="*80)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print_error("\n程序被用户中断，已保存当前进度，下次运行脚本可恢复！")
        sys.exit(0)
    except Exception as e:
        print_error(f"\n运行过程中出现错误：{str(e)}")
        print_help("如果是依赖安装失败，请手动运行：pip3 install pandas openpyxl matplotlib seaborn")
        print_help("如果是路径错误，请确保输入的路径正确，或切换到Rtab文件所在目录运行脚本")
        sys.exit(1)
